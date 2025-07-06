import requests, gzip, io
import pandas as pd
import numpy as np
import warnings
from openpyxl.styles import PatternFill

def process_data(
    product_codes_excel_path: str,
    output_excel_path: str,
    decl_codes_path: str = "https://raw.githubusercontent.com/tolgasabanoglu/eurostat_trade/main/data/ESTAT_CXT_PRODCOM_GEO.tsv",
    indicators_keep=None,
    years_to_drop=None,
    imputation: str = None,  # 'mean', 'median', 'mode'
    lint: bool = False  # Enable linear interpolation
):
    if indicators_keep is None:
        indicators_keep = ['EXPQNT', 'EXPVAL', 'IMPQNT', 'IMPVAL', 'PRODQNT', 'PRODVAL']
    if years_to_drop is None:
        years_to_drop = [str(year) for year in range(1995, 2018)]

    # Load product codes
    codes_df = pd.read_excel(product_codes_excel_path, dtype=str)
    codes_df.columns = ['prccode', 'description']

    # Download Comext data
    url = "https://ec.europa.eu/eurostat/api/comext/dissemination/sdmx/2.1/data/ds-056120/?format=TSV&compressed=true"
    r = requests.get(url)
    r.raise_for_status()
    data = gzip.decompress(r.content).decode('utf-8')
    df = pd.read_csv(io.StringIO(data), sep='\t')

    # Extract dimension columns
    df = df.rename(columns={df.columns[0]: "combined"})
    df[['freq', 'decl', 'prccode', 'indicators']] = df['combined'].str.split(',', expand=True)

    # Filter relevant product codes and indicators
    df_filtered = df[df['prccode'].isin(codes_df['prccode'])].copy()
    df_filtered = df_filtered[df_filtered['indicators'].isin(indicators_keep)]
    df_filtered.columns = df_filtered.columns.str.strip()

    # Drop unwanted year columns
    cols_to_drop = [col for col in years_to_drop if col in df_filtered.columns]
    df_filtered = df_filtered.drop(columns=cols_to_drop)

    # Identify year columns
    year_columns = sorted([col for col in df_filtered.columns if col.isdigit()])
    df_filtered[year_columns] = df_filtered[year_columns].replace(r'^\s*:?\s*$', np.nan, regex=True)
    df_filtered[year_columns] = df_filtered[year_columns].astype(float)

    # Save missing mask before any filling
    missing_mask = df_filtered[year_columns].isna()

    # Map decl codes to country names
    if decl_codes_path.startswith("http"):
        decl_codes = pd.read_csv(decl_codes_path, sep='\t', dtype=str)
    else:
        decl_codes = pd.read_csv(decl_codes_path, sep='\t', dtype=str)

    decl_codes.rename(columns={decl_codes.columns[0]: 'decl'}, inplace=True)
    mapping = decl_codes.set_index('decl')['Label - English'].to_dict()
    df_filtered['country'] = df_filtered['decl'].map(mapping)

    # -------------------- LINEAR INTERPOLATION --------------------
    if lint:
        missing_before_interp = df_filtered[year_columns].isna()

        def interpolate_single_gaps(row):
            vals = row.values
            n = len(vals)
            result = vals.copy()
            for i in range(1, n-1):
                if pd.isna(vals[i]):
                    if not pd.isna(vals[i-1]) and not pd.isna(vals[i+1]):
                        result[i] = (vals[i-1] + vals[i+1]) / 2
            return pd.Series(result, index=row.index)

        df_filtered[year_columns] = (
            df_filtered
            .groupby(['prccode', 'indicators'])[year_columns]
            .apply(lambda group: group.apply(interpolate_single_gaps, axis=1))
            .reset_index(level=[0, 1], drop=True)
        )

        interpolated_mask = missing_before_interp & (~df_filtered[year_columns].isna())
    else:
        interpolated_mask = pd.DataFrame(False, index=df_filtered.index, columns=year_columns)

    # -------------------- GROUP-BASED IMPUTATION --------------------
    if imputation:
        def impute_group(df_group):
            mask_zero = (df_group == 0).any(axis=1)
            mask_all_nan = df_group.isna().all(axis=1)
            skip_rows = mask_zero | mask_all_nan

            result = df_group.copy()

            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=RuntimeWarning)

                for idx in df_group.index[~skip_rows]:
                    row = result.loc[idx]
                    group_vals = result.loc[~skip_rows]

                    if imputation == 'mean':
                        impute_vals = group_vals.mean()
                    elif imputation == 'median':
                        impute_vals = group_vals.median()
                    elif imputation == 'mode':
                        mode_vals = group_vals.mode()
                        impute_vals = mode_vals.iloc[0] if not mode_vals.empty else pd.Series(np.nan, index=df_group.columns)
                    else:
                        raise ValueError("Invalid imputation method. Choose from 'mean', 'median', or 'mode'.")

                    to_fill = row.isna() & ~interpolated_mask.loc[idx]
                    row[to_fill] = impute_vals[to_fill]
                    result.loc[idx] = row

            return result

        df_filtered[year_columns] = (
            df_filtered
            .groupby(['prccode', 'indicators'])[year_columns]
            .apply(impute_group)
            .reset_index(level=[0, 1], drop=True)
        )

    # -------------------- MERGE DESCRIPTIONS --------------------
    df_filtered = df_filtered.merge(codes_df, on='prccode', how='left')
    if 'combined' in df_filtered.columns:
        df_filtered.drop(columns='combined', inplace=True)

    ordered_cols = ['country', 'indicators', 'description', 'prccode', 'freq', 'decl'] + year_columns
    df_filtered = df_filtered[ordered_cols]

    # -------------------- NETVOLQNT & NETVOLVAL Calculation --------------------
    def calc_netvol(df, value_type):
        relevant_inds = {
            'NETVOLVAL': ['PRODVAL', 'IMPVAL', 'EXPVAL'],
            'NETVOLQNT': ['PRODQNT', 'IMPQNT', 'EXPQNT'],
        }[value_type]

        all_rows = []

        for (country, prccode), group in df[df['indicators'].isin(relevant_inds)].groupby(['country', 'prccode']):
            pivoted = group.pivot(index='indicators', columns='year', values='value')
            if all(ind in pivoted.index for ind in relevant_inds):
                netvol = pivoted.loc[relevant_inds[0]] + pivoted.loc[relevant_inds[1]] - pivoted.loc[relevant_inds[2]]
                netvol.name = (country, prccode, value_type)
                all_rows.append(netvol)

        df_netvol = pd.DataFrame(all_rows)
        df_netvol.reset_index(inplace=True)
        df_netvol.columns = ['country', 'prccode', 'indicators'] + list(df_netvol.columns[3:])
        return df_netvol

    df_long = df_filtered.melt(
        id_vars=['country', 'prccode', 'indicators'],
        value_vars=year_columns,
        var_name='year',
        value_name='value'
    )

    df_netval = calc_netvol(df_long, 'NETVOLVAL')
    df_netqnt = calc_netvol(df_long, 'NETVOLQNT')
    netvol_final = pd.concat([df_netval, df_netqnt], ignore_index=True)

    netvol_final = netvol_final.merge(codes_df, on='prccode', how='left')
    netvol_final['freq'] = 'A'
    netvol_final['decl'] = np.nan
    netvol_final = netvol_final[['country', 'indicators', 'description', 'prccode', 'freq', 'decl'] + year_columns]

    # -------------------- SAVE TO EXCEL --------------------
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_filtered.to_excel(writer, index=False, sheet_name='Data')
        workbook = writer.book
        worksheet = writer.sheets['Data']

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        for row_idx, (miss_row, interp_row) in enumerate(zip(missing_mask.itertuples(index=False),
                                                             interpolated_mask.itertuples(index=False)), start=2):
            for col_idx, (was_missing, was_interpolated) in enumerate(zip(miss_row, interp_row), start=7):
                if was_interpolated:
                    worksheet.cell(row=row_idx, column=col_idx).fill = blue_fill
                elif was_missing:
                    worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

        netvol_final.to_excel(writer, index=False, sheet_name='NetVolumes')

    print(f"✅ Processed data saved to: {output_excel_path}")
    return df_filtered
